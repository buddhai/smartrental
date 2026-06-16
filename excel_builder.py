# -*- coding: utf-8 -*-
"""
엑셀 수익률분석표 템플릿을 렌탈 기간(개월)에 맞게 조정합니다.
"""
import re
from copy import copy

import openpyxl
from openpyxl.styles import Border, Side

CALC_SHEET_IDX = 1
CHECK_SHEET_IDX = 0
BASE_TERM = 12
CALC_FIRST_PAY = 14
CALC_ROW0 = 13
CALC_LAST_PAY_BASE = 25
CALC_TERMINAL_BASE = 26
CHECK_FIRST_PAY = 11
CHECK_LAST_PAY_BASE = 22
STYLE_TEMPLATE_ROW = 15  # 행 확장 시 스타일 복사 기준 (25행은 테두리 없음)

_THIN = Side(style='thin')
_FULL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _copy_cell_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = copy(src.number_format)
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)


def _calc_terminal_row(term: int) -> int:
    return CALC_FIRST_PAY + term


def _apply_calc_row_borders(ws, row: int):
    """현금흐름표 데이터 행에 테두리 적용 (선수금 K열 포함)."""
    for col in range(2, 17):  # B~P
        cell = ws.cell(row, col)
        cell.border = _FULL_BORDER


def _set_calc_payment_row(ws, row: int, period: int, terminal_row: int, is_first: bool):
    ws[f'B{row}'] = period
    if is_first:
        ws[f'C{row}'] = '=K6'
        ws[f'E{row}'] = '=-E9*F6'
        ws[f'G{row}'] = '=-F9'
        ws[f'H{row}'] = '=-G9'
        ws[f'M{row}'] = f'=(G6-J{terminal_row})/H6'
        ws[f'N{row}'] = '=N13-M14'
        ws[f'O{row}'] = '=(N13*$N$6)/12'
        ws[f'P{row}'] = f'=L{row}-M{row}-O{row}'
    else:
        prev = row - 1
        ws[f'C{row}'] = f'=C{prev}'
        if row == CALC_FIRST_PAY + 1:
            ws[f'E{row}'] = ' '
        else:
            ws[f'E{row}'] = None
        ws[f'G{row}'] = f'=G{prev}'
        ws[f'H{row}'] = f'=H{prev}'
        ws[f'M{row}'] = f'=M{prev}'
        ws[f'N{row}'] = f'=N{prev}-M{row}'
        ws[f'O{row}'] = f'=(N{prev}*$N$6)/12'
        ws[f'P{row}'] = f'=L{row}-M{row}-O{row}'
    ws[f'L{row}'] = f'=SUM(C{row}:K{row})'
    _apply_calc_row_borders(ws, row)


def _set_calc_terminal_row(ws, row: int, term: int):
    ws[f'B{row}'] = term + 1
    ws[f'L{row}'] = f'=SUM(C{row}:K{row})'
    ws[f'P{row}'] = f'=L{row}-M{row}-O{row}'
    _apply_calc_row_borders(ws, row)


def resize_calc_sheet(ws, term: int):
    if term < 1:
        raise ValueError('렌탈 기간은 1개월 이상이어야 합니다.')

    terminal_row = _calc_terminal_row(term)
    delta = term - BASE_TERM

    if delta > 0:
        ws.insert_rows(CALC_TERMINAL_BASE, delta)
        for i in range(delta):
            new_row = CALC_TERMINAL_BASE + i
            for col in range(1, 17):
                src = ws.cell(STYLE_TEMPLATE_ROW, col)
                dst = ws.cell(new_row, col)
                _copy_cell_style(src, dst)

    elif delta < 0:
        ws.delete_rows(CALC_FIRST_PAY + term, -delta)

    for p in range(1, term + 1):
        row = CALC_FIRST_PAY + p - 1
        _set_calc_payment_row(ws, row, p, terminal_row, is_first=(p == 1))

    terminal_row = _calc_terminal_row(term)
    _set_calc_terminal_row(ws, terminal_row, term)
    _apply_calc_row_borders(ws, CALC_ROW0)

    ws['L12'] = f'=IRR(L{CALC_ROW0}:L{terminal_row})*12'

    sum_row = terminal_row + 1
    ws[f'P{sum_row}'] = f'=SUM(P{CALC_FIRST_PAY}:P{terminal_row})'


def _set_check_payment_row(ws, row: int, period: int, is_first: bool):
    ws[f'A{row}'] = period
    if is_first:
        ws[f'B{row}'] = '=E4'
        ws[f'C{row}'] = f'=B{row}/$B$3'
        ws[f'D{row}'] = 0
        ws[f'F{row}'] = f'=B{row}+E{row}'
        ws[f'G{row}'] = f'=F{row}/$B$3'
        ws[f'H{row}'] = '=$E$5-B11'
        ws[f'I{row}'] = '=+$H$4'
        ws[f'J{row}'] = '=$H$6'
        ws[f'K{row}'] = f'=F{row}+I{row}+J{row}'
        ws[f'L{row}'] = f'=K{row}/$B$3'
        ws[f'M{row}'] = '=$B$3-K11'
    else:
        prev = row - 1
        ws[f'B{row}'] = f'=$B$11*A{row}'
        ws[f'C{row}'] = f'=B{row}/$B$3'
        ws[f'D{row}'] = 0
        ws[f'F{row}'] = f'=B{row}+E{row}'
        ws[f'G{row}'] = f'=F{row}/$B$3'
        ws[f'H{row}'] = f'=$E$5-B{row}'
        ws[f'I{row}'] = f'=IF(H{row}>I{prev},I{prev},H{row})'
        ws[f'J{row}'] = '=$H$6'
        ws[f'K{row}'] = f'=F{row}+I{row}+J{row}'
        ws[f'L{row}'] = f'=K{row}/$B$3'
        ws[f'M{row}'] = f'=$B$3-K{row}'


def resize_check_sheet(ws_calc, ws_check, term: int):
    calc_name = ws_calc.title
    terminal_row = _calc_terminal_row(term)
    ws_check['B4'] = f"=+'{calc_name}'!H6"
    ws_check['E7'] = f"='{calc_name}'!J{terminal_row}"

    delta = term - BASE_TERM
    if delta > 0:
        ws_check.insert_rows(CHECK_LAST_PAY_BASE + 1, delta)
        template_row = CHECK_FIRST_PAY + 1
        for i in range(delta):
            new_row = CHECK_LAST_PAY_BASE + 1 + i
            for col in range(1, 14):
                src = ws_check.cell(template_row, col)
                dst = ws_check.cell(new_row, col)
                _copy_cell_style(src, dst)

    elif delta < 0:
        ws_check.delete_rows(CHECK_FIRST_PAY + term, -delta)

    for p in range(1, term + 1):
        row = CHECK_FIRST_PAY + p - 1
        _set_check_payment_row(ws_check, row, p, is_first=(p == 1))


def _replace_sheet_ref(formula, old_name: str, new_name: str):
    if not isinstance(formula, str):
        return formula
    # '시그마크 12개월'!A1 형태 치환
    return formula.replace(f"'{old_name}'", f"'{new_name}'")


def rename_workbook_sheets(wb: openpyxl.Workbook, term: int):
    """시트명을 개월수 기준으로 변경하고 체크시트 참조 수식 갱신."""
    ws_calc = wb.worksheets[CALC_SHEET_IDX]
    ws_check = wb.worksheets[CHECK_SHEET_IDX]
    old_calc = ws_calc.title
    old_check = ws_check.title
    new_calc = f'수익률_{term}개월'[:31]
    new_check = f'체크_{term}개월'[:31]

    if old_calc != new_calc:
        for row in ws_check.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    cell.value = _replace_sheet_ref(cell.value, old_calc, new_calc)
        ws_calc.title = new_calc

    if old_check != new_check:
        ws_check.title = new_check


def set_sga_rate(ws, term: int, sga_rate_pct: float):
    """
    HTML 판관비율(%)을 엑셀 I9에 반영.
    엑셀: F9=(G6*I9)/H6 → 월 판관비 = G6 * (sga_rate_pct/100)
    """
    monthly_decimal = float(sga_rate_pct) / 100.0
    ws['I9'] = monthly_decimal * int(term)


def fill_input_cells(ws, data: dict):
    """HTML 입력값을 계산 시트에 반영."""
    product = data.get('productName', '')
    term = int(data.get('term', 12))
    cost = float(data.get('cost', 0))
    qty = int(data.get('qty', 1))
    total_fee = float(data.get('totalMonthlyFee', 0))
    advance = float(data.get('advance', 0))
    borrow_rate = float(data.get('borrowRate', 0)) / 100
    residual = float(data.get('residual', 0))
    sga_rate = float(data.get('sgaRate', 0.105555555555))
    irr_type = data.get('irrType', 'unlevered')

    ws['B2'] = product
    ws['D6'] = product
    ws['E6'] = cost
    ws['F6'] = qty
    ws['H6'] = term
    ws['K6'] = total_fee
    ws['K14'] = advance if advance else None
    ws['N6'] = 0 if irr_type == 'unlevered' else borrow_rate
    ws['O6'] = residual
    set_sga_rate(ws, term, sga_rate)


def prepare_workbook(wb: openpyxl.Workbook, term: int) -> openpyxl.Workbook:
    term = int(term)
    ws_calc = wb.worksheets[CALC_SHEET_IDX]
    ws_check = wb.worksheets[CHECK_SHEET_IDX]

    resize_calc_sheet(ws_calc, term)
    resize_check_sheet(ws_calc, ws_check, term)
    rename_workbook_sheets(wb, term)

    ws_calc['H6'] = term
    return wb


_INVALID_SHEET_CHARS = re.compile(r'[\\/*?:\[\]]')


def _sanitize_sheet_token(name: str) -> str:
    name = _INVALID_SHEET_CHARS.sub('', str(name or '')).strip()
    return re.sub(r'\s+', '', name)


def _unique_sheet_name(base: str, used: set) -> str:
    base = (base or '시나리오')[:31]
    if base not in used:
        used.add(base)
        return base
    for i in range(2, 100):
        suffix = f'_{i}'
        candidate = f'{base[:31 - len(suffix)]}{suffix}'
        if candidate not in used:
            used.add(candidate)
            return candidate
    raise ValueError('시트 이름을 더 이상 만들 수 없습니다.')


def allocate_scenario_sheet_names(product: str, term: int, used: set) -> tuple:
    """시나리오별 계산/체크 시트명 생성."""
    token = _sanitize_sheet_token(product)[:18] or '상품'
    calc_name = _unique_sheet_name(f'{token}_{term}M', used)
    check_name = _unique_sheet_name(f'체크_{calc_name}'[:31], used)
    return calc_name, check_name


def rename_scenario_pair(ws_calc, ws_check, calc_name: str, check_name: str):
    """계산·체크 시트 쌍의 이름과 상호 참조 수식 갱신."""
    calc_name = calc_name[:31]
    check_name = check_name[:31]
    old_calc = ws_calc.title

    for row in ws_check.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                cell.value = _replace_sheet_ref(cell.value, old_calc, calc_name)

    ws_calc.title = calc_name
    ws_check.title = check_name


def copy_worksheet(src_ws, dest_wb: openpyxl.Workbook, title: str):
    """워크북 간 시트 복사 (서식·병합 포함)."""
    title = title[:31]
    if title in dest_wb.sheetnames:
        raise ValueError(f'시트 이름 충돌: {title}')

    dest_ws = dest_wb.create_sheet(title=title)
    for row in src_ws.iter_rows():
        for cell in row:
            dst = dest_ws.cell(cell.row, cell.column, cell.value)
            _copy_cell_style(cell, dst)

    for col, dim in src_ws.column_dimensions.items():
        dest_ws.column_dimensions[col].width = dim.width
    for row, dim in src_ws.row_dimensions.items():
        dest_ws.row_dimensions[row].height = dim.height
    for merged in list(src_ws.merged_cells.ranges):
        dest_ws.merge_cells(str(merged))

    dest_ws.sheet_format = copy(src_ws.sheet_format)
    dest_ws.sheet_properties = copy(src_ws.sheet_properties)
    dest_ws.page_setup = copy(src_ws.page_setup)
    dest_ws.print_options = copy(src_ws.print_options)
    return dest_ws


def _merge_line_with_globals(line: dict, globals_: dict) -> dict:
    merged = dict(globals_)
    merged.update({k: v for k, v in line.items() if v is not None and v != ''})
    return merged


def add_summary_sheet(wb: openpyxl.Workbook, scenarios: list):
    """협의용 요약 시트 (맨 앞)."""
    ws = wb.create_sheet('요약', 0)
    headers = ['No', '상품', '기간(월)', '수량', '취득가', '월렌탈료', 'IRR(%)', '수익률시트', '체크시트']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.border = _FULL_BORDER

    for i, sc in enumerate(scenarios, 1):
        row = i + 1
        calc = sc['calc_sheet']
        check = sc['check_sheet']
        ws.cell(row, 1, i)
        ws.cell(row, 2, sc.get('productName', ''))
        ws.cell(row, 3, sc.get('term', ''))
        ws.cell(row, 4, sc.get('qty', ''))
        ws.cell(row, 5, sc.get('cost', ''))
        ws.cell(row, 6, sc.get('totalMonthlyFee', ''))
        ws.cell(row, 7, f"='{calc}'!L12")
        ws.cell(row, 8, calc)
        ws.cell(row, 9, check)
        for col in range(1, 10):
            ws.cell(row, col).border = _FULL_BORDER

    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18


def build_bundle_workbook(lines: list, globals_: dict, load_workbook_for_term) -> openpyxl.Workbook:
    """
    시나리오 목록으로 단일 워크북 생성.
    load_workbook_for_term(term) -> openpyxl.Workbook 콜백.
    """
    if not lines:
        raise ValueError('시나리오가 없습니다.')

    bundle = openpyxl.Workbook()
    bundle.remove(bundle.active)
    used_names = set()
    scenario_meta = []

    for line in lines:
        data = _merge_line_with_globals(line, globals_)
        term = int(data.get('term', 12))
        product = data.get('productName', '') or '미지정'

        wb = load_workbook_for_term(term)
        ws_calc = wb.worksheets[CALC_SHEET_IDX]
        ws_check = wb.worksheets[CHECK_SHEET_IDX]

        fill_input_cells(ws_calc, data)
        calc_name, check_name = allocate_scenario_sheet_names(product, term, used_names)
        rename_scenario_pair(ws_calc, ws_check, calc_name, check_name)

        copy_worksheet(ws_check, bundle, check_name)
        copy_worksheet(ws_calc, bundle, calc_name)

        scenario_meta.append({
            **data,
            'calc_sheet': calc_name,
            'check_sheet': check_name,
        })

    add_summary_sheet(bundle, scenario_meta)
    return bundle
