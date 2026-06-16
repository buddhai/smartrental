# -*- coding: utf-8 -*-
"""개월수별 템플릿 생성 (선택 실행)"""
import os
import openpyxl
from excel_builder import prepare_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATES_DIR = os.path.join(BASE_DIR, 'templates')
BASE_PATH = os.path.join(TEMPLATES_DIR, 'base.xlsx')
TERMS = [12, 24, 36, 48, 60]


def generate_all():
    os.makedirs(TEMPLATES_DIR, exist_ok=True)
    for term in TERMS:
        wb = openpyxl.load_workbook(BASE_PATH)
        prepare_workbook(wb, term)
        out = os.path.join(TEMPLATES_DIR, f'수익률분석표_{term}개월.xlsx')
        wb.save(out)
        ws = wb.worksheets[1]
        terminal = 14 + term
        print(f'생성: {term}개월 -> {out}')
        print(f'  IRR수식: {ws["L12"].value}')
        print(f'  마지막회차 B{terminal-1}={ws[f"B{terminal-1}"].value}, 터미널 B{terminal}={ws[f"B{terminal}"].value}')


if __name__ == '__main__':
    generate_all()
