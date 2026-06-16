# -*- coding: utf-8 -*-
"""
수익률분석표 자동완성 서버
실행: python server.py
접속: http://localhost:5000
"""
import os
import io
import glob
from datetime import datetime

import openpyxl
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS

from excel_builder import prepare_workbook, fill_input_cells, rename_workbook_sheets, build_bundle_workbook

app = Flask(__name__)
CORS(app)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATES_DIR = os.path.join(BASE_DIR, 'templates')
SUPPORTED_TERMS = {12, 24, 36, 48, 60}


def find_base_template():
    candidates = [os.path.join(TEMPLATES_DIR, 'base.xlsx')]
    for p in candidates:
        if os.path.exists(p):
            return p
    for p in glob.glob(os.path.join(BASE_DIR, '수익률분석표*.xlsx')):
        return p
    return None


def get_template_path(term: int):
    cached = os.path.join(TEMPLATES_DIR, f'수익률분석표_{term}개월.xlsx')
    if os.path.exists(cached):
        return cached
    return find_base_template()


def load_workbook_for_term(term: int):
    term = int(term)
    path = get_template_path(term)
    if not path:
        raise FileNotFoundError('템플릿 파일을 찾을 수 없습니다.')

    wb = openpyxl.load_workbook(path)
    cached = path.endswith(f'_{term}개월.xlsx')
    if not cached:
        prepare_workbook(wb, term)
    else:
        rename_workbook_sheets(wb, term)
    return wb


@app.route('/')
def index():
    return send_file(os.path.join(BASE_DIR, '렌탈 견적 산정시스템_v2.4.html'))


def _validate_term(term: int):
    if term not in SUPPORTED_TERMS:
        raise ValueError(
            f'렌탈 기간 {term}개월은 지원하지 않습니다. '
            f'지원: {", ".join(str(t) for t in sorted(SUPPORTED_TERMS))}개월'
        )


def _parse_line(data: dict, globals_: dict = None) -> dict:
    g = globals_ or {}
    merged = {**g, **data}
    term = int(merged.get('term', g.get('term', 12)))
    _validate_term(term)
    product = (merged.get('productName') or '').strip()
    if not product:
        raise ValueError('상품명이 비어 있습니다.')
    return merged


@app.route('/fill-excel-bundle', methods=['POST'])
def fill_excel_bundle():
    try:
        body = request.get_json() or {}
        lines = body.get('lines') or []
        if not lines:
            return jsonify({'error': '시나리오(lines)가 없습니다.'}), 400

        globals_ = {
            'borrowRate': body.get('borrowRate', 6),
            'sgaRate': body.get('sgaRate', 0.105555555555),
            'irrType': body.get('irrType', 'unlevered'),
            'residual': body.get('residual', 0),
        }

        parsed_lines = []
        for i, line in enumerate(lines, 1):
            try:
                parsed_lines.append(_parse_line(line, globals_))
            except ValueError as e:
                return jsonify({'error': f'{i}번 시나리오: {e}'}), 400

        wb = build_bundle_workbook(parsed_lines, globals_, load_workbook_for_term)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        date_str = datetime.now().strftime('%Y%m%d')
        filename = f'수익률분석_협의_{len(parsed_lines)}건_{date_str}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    except FileNotFoundError as e:
        return jsonify({'error': str(e)}), 404
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/fill-excel', methods=['POST'])
def fill_excel():
    try:
        data = request.get_json() or {}
        term = int(data.get('term', 12))
        _validate_term(term)

        wb = load_workbook_for_term(term)
        ws = wb.worksheets[1]
        fill_input_cells(ws, data)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        product_name = data.get('productName', '')
        date_str = datetime.now().strftime('%Y%m%d')
        filename = f'수익률분석표_{product_name}_{term}개월_{date_str}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    except FileNotFoundError as e:
        return jsonify({'error': str(e)}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    print('=' * 50)
    print('수익률분석표 서버 시작')
    print('지원 렌탈기간:', ', '.join(f'{t}개월' for t in sorted(SUPPORTED_TERMS)))
    print('브라우저에서 http://localhost:5000 접속')
    print('=' * 50)
    app.run(host='0.0.0.0', port=5000, debug=False)
